import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill

with analytics.track():
# --- PAGE CONFIG ---
st.set_page_config(page_title="Name Validator", page_icon="✅")

st.title("User name Validator")
st.markdown("""
Upload your Excel file to verify names. 
**Logic:** At least **one** name must match 100% between the columns. This is a testing script, don't forget to manually confirm and validate.  
""")

# --- FILE UPLOAD ---
uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx'])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    
    # Column Selection (Flexible in case names change)
    col1, col2 = st.columns(2)
    with col1:
        name_col = st.selectbox("Select 'Customer Name' column", df.columns)
    with col2:
        error_col = st.selectbox("Select 'Validation/Error' column", df.columns)

    if st.button("🚀 Process and Validate"):
        
        def exact_match_logic(row):
            # 1. Clean and tokenize
            set1 = set(str(row[name_col]).strip().lower().split())
            set2 = set(str(row[error_col]).strip().lower().split())
            
            # 2. Check for any exact intersection
            common = set1.intersection(set2)
            
            status = "VERIFIED" if len(common) >= 1 else "FLAGGED"
            return pd.Series([len(common), status], index=['Match Count', 'System Status'])

        # Apply Logic
        df[['Match Count', 'System Status']] = df.apply(exact_match_logic, axis=1)

        # Create Styled Excel in Memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            
            # Styling
            workbook = writer.book
            worksheet = writer.sheets['Results']
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            status_col_idx = df.columns.get_loc("System Status") + 1
            for row_idx in range(2, len(df) + 2):
                if worksheet.cell(row=row_idx, column=status_col_idx).value == "FLAGGED":
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

        # --- DOWNLOAD BUTTON ---
        st.success(f"Processing Complete! Found {len(df[df['System Status']=='FLAGGED'])} discrepancies.")
        
        st.download_button(
            label="📥 Download Flagged Report",
            data=output.getvalue(),
            file_name="Validated_Identity_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )